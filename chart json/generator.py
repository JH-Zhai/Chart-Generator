import os
import openpyxl
import json
import sys
import subprocess

"""
This project was created by Zhai Jia Hong. https://github.com/JH-Zhai

Feel free to contribute to this project.

This project is open-source and free to use. You are welcome to modify and distribute it.

GLHF
"""

# Get the directory of the script
dir = os.path.dirname(os.path.abspath(sys.argv[0]))
print("directory:", dir)

# Construct the path to the Excel file relative to the script directory
excel_file = os.path.join(dir, 'data.xlsx')
# print("Excel file path:", excel_file)

# Check if the Excel file exists
if not os.path.exists(excel_file):
    print("Excel file does not exist.")
    sys.exit(1)

# html_file = os.path.join(script_dir, 'test.html')
# print("Html file path:", html_file)

# # Check if the Html file exists
# if not os.path.exists(html_file):
#     print("Html file does not exist.")
#     sys.exit(1)

# # Open the HTML file in the default web browser
# webbrowser.open('test.html')

# Load the workbook
workbook = openpyxl.load_workbook(excel_file, read_only=True)

# Get the node and edge sheets
node_sheet = workbook['node']
edge_sheet = workbook['edge']

# Read data from node sheet
nodes = []
for row in node_sheet.iter_rows(min_row=2, values_only=True):
    node = {
        "label": row[0],
        "color": row[1],
        "y": row[2],
        "x": row[3],
        "id": row[4],
        "size": row[5],
        "shape": row[6]
    }
    if row[7]:
        node["note"] = row[7]
    nodes.append(node)

# Read data from edge sheet
edges = []
for row in edge_sheet.iter_rows(min_row=2, values_only=True):
    edge = {
        "sourceID": row[0],
        "targetID": row[1],
        "size": row[2]
    }
    if row[3]:
        edge["note"] = row[3]
    edges.append(edge)

# Combine nodes and edges into a dictionary
graph_json = {'nodes': nodes, 'edges': edges}

# Construct the path to the JSON file relative to the script directory
json_file = os.path.join(dir, 'data.json')
# print("JSON file path:", json_file)

# Write the JSON to a file
with open(json_file, 'w') as f:
    json.dump(graph_json, f, ensure_ascii=False, indent=3)

# print("JSON file successfully created.")

# Command to open the HTML file in Google Chrome
command = 'open -a "Safari" ' + dir + '/graph.html'

# print(command)

# Execute the command
subprocess.call(command, shell=True)

# input("Press Enter to exit...")