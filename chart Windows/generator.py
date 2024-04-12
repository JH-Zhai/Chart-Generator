"""
This project was created by Zhai Jia Hong. https://github.com/JH-Zhai

Feel free to contribute to this project.

This project is open-source and free to use. You are welcome to modify and distribute it.

GLHF
"""

import os
import openpyxl
import sys
import subprocess

# Get the directory of the script
dir = os.path.dirname(os.path.abspath(sys.argv[0]))
print("directory:", dir)


# Construct the path to the Excel file relative to the script directory
excel_file = os.path.join(dir, 'data.xlsx')
# print("Excel file path:", excel_file)


# Check if the Excel file exists
if not os.path.exists(excel_file):
    print("Excel file does not exist.")
    sys.exit(1)


# Load the workbook
workbook = openpyxl.load_workbook(excel_file, read_only=True)


# Get the node and edge sheets
node_sheet = workbook['node']
edge_sheet = workbook['edge']


# Read data from node sheet
nodes = []
for row in node_sheet.iter_rows(min_row=2, values_only=True):
    node = {
        "label": row[0],
        "color": row[1],
        "y": row[2],
        "x": row[3],
        "id": row[4],
        "size": row[5],
        "shape": row[6]
    }
    if row[7]:
        node["note"] = row[7]
    nodes.append(node)


# Read data from edge sheet
edges = []
for row in edge_sheet.iter_rows(min_row=2, values_only=True):
    edge = {
        "sourceID": row[0],
        "targetID": row[1],
        "size": row[2]
    }
    if row[3]:
        edge["note"] = row[3]
    edges.append(edge)






dataString = "<!DOCTYPE html>"


dataString += """
<html lang="en" style="height: 100%">


<head>
    <!-- <meta charset="utf-8"> -->
    <title>Blooming</title>
</head>


<body style="height: 100%; margin: 0">
    <div id="container" style="height: 100%"></div>
    <script type="text/javascript" src="https://cdn.staticfile.net/jquery/3.7.1/jquery.min.js"></script>
    <script type="text/javascript"
        src="https://registry.npmmirror.com/echarts/5.5.0/files/dist/echarts.min.js"></script>
    <script type="text/javascript">
   
        // Initializing ECharts instance
        var dom = document.getElementById('container');
       
        var myChart = echarts.init(dom, null, {
            renderer: 'canvas',
            useDirtyRect: false
        });
       
        var option;


        var jsonData = {


"""


dataString += '            "nodes": [ \n'


for node in nodes:
    dataString += '                \n'
    dataString += '                {\n'
   
    dataString += '                    "label": "'
    dataString += node["label"]
    dataString += '", \n'


    dataString += '                    "color": "'
    dataString += node["color"]
    dataString += '", \n'


    dataString += '                    "y": '
    dataString += str(node["y"])
    dataString += ', \n'


    dataString += '                    "x": '
    dataString += str(node["x"])
    dataString += ', \n'


    dataString += '                    "id": "'
    dataString += node["id"]
    dataString += '", \n'


    dataString += '                    "size": '
    dataString += str(node["size"])
    dataString += ', \n'


    dataString += '                    "shape": "'
    dataString += node["shape"]
    dataString += '"'


    if("note" in node):
        dataString += ', \n                    "note": "'
        dataString += node["note"]
        dataString += '"'


    dataString += '\n                },'
    dataString += '\n'


dataString += '            ],\n \n'


dataString += '            "edges": [ \n'


for edge in edges:
    dataString += '                \n'
    dataString += '                {\n'
   
    dataString += '                    "sourceID": "'
    dataString += edge["sourceID"]
    dataString += '", \n'
   
    dataString += '                    "targetID": "'
    dataString += edge["targetID"]
    dataString += '", \n'
   
    dataString += '                    "size": '
    dataString += str(edge["size"])
   
    if("note" in edge):
        dataString += ', \n                    "note": "'
        dataString += edge["note"]
        dataString += '"'
   
    dataString += '\n                },'
    dataString += '\n'
   
dataString += '            ]\n'




dataString += """
        }




        // Show loading animation while fetching data
        myChart.showLoading();
        myChart.hideLoading();
        myChart.setOption({
            title: {
                text: '物质关系'
            },
            animationDurationUpdate: 1500,
            animationEasingUpdate: 'quinticInOut',
            series: [
                {
                    type: 'graph',
                    layout: 'none',
                    roam: true, // Enable mouse zoom and pan
                    draggable: true, // Enable node dragging




                    data: jsonData.nodes.map(function (node) {
                        return {
                            x: node.x,
                            y: node.y,
                            id: node.id,
                            name: node.label,
                            note: node.note,
                            symbolSize: node.size,
                            symbol: node.shape,
                            itemStyle: { color: node.color },
                            label: {
                                show: true, // Always show labels
                                formatter: function (params) {
                                    // Concatenate node's name and note
                                    if (params.data.note === undefined) {
                                        return params.data.name
                                    } else {
                                        return params.data.name + '\\n' + params.data.note
                                    }
                                },
                                position: 'inside', // Adjust label position as needed
                                fontSize: 12
                            },
                        }
                    }),




                    edges: jsonData.edges.map(function (edge) {
                        if (edge.note === undefined) {
                            return {
                                source: edge.sourceID,
                                target: edge.targetID,
                                lineStyle: {
                                    width: 2 // Adjust thickness of the edge
                                },
                                symbol: ['none', 'arrow'], // Show arrow at target
                                symbolSize: 10,
                                lineStyle: {
                                    width: 3,
                                    curveness: 0,
                                    opacity: 0.7
                                }// Size of the arrow
                            };
                        } else {
                            return {
                                source: edge.sourceID,
                                target: edge.targetID,
                                lineStyle: {
                                    width: 2 // Adjust thickness of the edge
                                },
                                symbol: ['none', 'arrow'], // Show arrow at target
                                symbolSize: 10,
                                label: {
                                    show: true, // Show label for edge
                                    formatter: edge.note, // Display note as label
                                    position: 'middle', // Adjust label position
                                    fontSize: 10,
                                    color: '#333' // Adjust label color
                                },
                                lineStyle: {
                                    width: 3,
                                    curveness: 0,
                                    opacity: 0.7
                                }// Size of the arrow
                            };
                        }
                    }),
                    emphasis: {
                        focus: 'adjacency'
                    },
                }
            ]
        }, true)
        // Ensure option object is defined and set it to the chart
        if (option && typeof option === 'object') {
            myChart.setOption(option);
        }




        // Resize chart when window is resized
        window.addEventListener('resize', myChart.resize);
    </script>
</body>




</html>


"""




# print(dataString)
html_file = os.path.join(dir, 'graph.html')


with open(html_file, "w") as f:
    f.write(dataString)
   
   
# # Combine nodes and edges into a dictionary
# graph_json = {'nodes': nodes, 'edges': edges}


# # Construct the path to the JSON file relative to the script directory
# json_file = os.path.join(dir, 'data.json')
# # print("JSON file path:", json_file)


# # Write the JSON to a file
# with open(json_file, 'w') as f:
#     json.dump(graph_json, f, ensure_ascii=False, indent=3)


# Command to open the HTML file in Google Chrome
command = f'start msedge "{html_file}"'


# Execute the command
subprocess.call(command, shell=True)


input("press enter to exit")